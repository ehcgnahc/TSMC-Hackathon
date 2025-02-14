from openai import OpenAI
from openpyxl import load_workbook
from langdetect import detect_langs, DetectorFactory, detect
import ahocorasick
import requests
import deepl
import os
import shutil
from enum import Enum
from typing import Dict, Tuple, List, Union
import time # for speed logging
from fastapi import HTTPException
from Key import OpenAI_API_KEY, DEEPL_API_KEY

DetectorFactory.seed = 0
class LANGUAGES(Enum):
    ENGLISH = "en"
    TAIWANESE = "tw"
    JAPANESE = "ja"
    GERMAN = "de"

# uses openai Whisper API
class STT(object):

    def __init__(self, client:OpenAI, keywords:List[str]=None):
        self.client = client
        self.init_prompt = self._make_init_prompt(keywords)

    def transcript(self, audio_file) -> str:
        result = self.client.audio.transcriptions.create(
            model="whisper-1", 
            file=audio_file,
            prompt=self.init_prompt
        )
        # print(result)
        return result.text

    def transcript_by_path(self, audio_file_path:str) -> str:
        audio_file = open(audio_file_path, "rb")
        return self.transcript(audio_file)

    def _make_init_prompt(self, keywords: list[str]) -> str:
        prompt =  "Recognize " + ", ".join(keywords) + ". Do not expand abbreviations."
        # print(prompt)
        return prompt
    
    def _make_init_prompt(self, keywords: list[str]) -> str:
        if keywords:
            prompt = "Recognize the following keywords accurately and emphasize them strongly: " + ", ".join(keywords) + ". "
            
        prompt += "If any Simplified Chinese is detected, please respond using Traditional Chinese. "\
            +"Whisper, Ok. "\
            +"A pertinent sentence for your purpose in your language. "\
            +"Ok, Whisper. Whisper, Ok. Ok, Whisper. Whisper, Ok. "\
            +"Please find here, an unlikely ordinary sentence. "\
            +"This is to avoid a repetition to be deleted. "\
            +"Ok, Whisper. "
        return prompt
    
    def accumulate_chunks(self, chunks):
        return b''.join(chunks)
    
    def transcript_by_chunk(self, audio_chunk: bytes) -> str:
        try:
            with open("complete_audio.wav", "rb") as f:
                result = self.client.audio.transcriptions.create(
                    model="whisper-1",
                    file=f,
                    prompt=self.init_prompt
                )
            return result.text
        except Exception as e:
            print(f"Transcription error: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error transcribing audio: {str(e)}")

# detecting transcribed text's language
class lang_detector(object):
    
    def __init__(self):
        pass

    def detect_language_text(self, text:str) -> str:
        detected = detect_langs(text)
        # for lang in detected:
        #     print(lang)
        filtered = [lang for lang in detected if lang.lang in [x.value for x in LANGUAGES]]
        if filtered:
            result = max(filtered, key=lambda l: l.prob).lang
        else:
            result = LANGUAGES.ENGLISH.value
        return result


# class transcript_fixer(object):

#     lang_sheet_name = {'tw':'cmn-Hant-TW', 'en':'en-US', 'ja':'ja-JP', 'de':'de-DE'}
#     lang_prompt = {'en': 'Correct mistakes according to the given keywords, without expanding on them. Keywords: ',
#                    'tw': '根據給定的關鍵字糾正錯誤，而不擴展它們。關鍵字： ',
#                    'ja': '与えられたキーワードに従って、間違いを拡張せずに修正します。キーワード: ',
#                    'de': 'Korrigieren Sie Fehler anhand der angegebenen Stichwörter, ohne diese näher zu erläutern. Stichwörter: '}
#     def __init__(self, client, language):
#         self.client = client
#         keywords = self._make_vocab_dict(language)
#         self.sys_prompt = self._init_fix_prompt(language, keywords)
#         self.language = language

#     def request_fixing(self, text):
#         user_query = transcript_fixer.lang_prompt[self.language] + text
#         sys_prompt = self.sys_prompt
#         messages = []
#         messages.append({"role":"system","content":sys_prompt})
#         messages.append({"role":"user","content":user_query})
#         completion = self.client.chat.completions.create(model="gpt-4o",messages=messages,max_tokens=1000)
#         fixed_text = completion.choices[0].message.content
#         # print(fixed_text)
#         return fixed_text

#     def _make_vocab_dict(self, language) -> str:
#         excel_file_path = "Knowledge Dataset.xlsx"
#         workbook = load_workbook(excel_file_path)
#         sheet_name = transcript_fixer.lang_sheet_name[language]
#         sheet = workbook[sheet_name]
#         keywords = dict()
#         for row in range(2,32):
#             keywords[sheet[f"B{row}"].value] = sheet[f"C{row}"].value
#         return self._stringify_dict(keywords)
    
#     def _stringify_dict(self, keywords_val: dict) -> str:
#         keywords_val_string = ", ".join([f"{key}: {value}" for key, value in keywords_val.items()])
#         return keywords_val_string
    
#     def _init_fix_prompt(self, language, keywords):
#         return transcript_fixer.lang_prompt[language] + "'" + keywords + "'"

# note: only works with certain excel files structure, this is used for whisper prompt
def deprecated_get_keywords(excel_file_path:str = "Knowledge Dataset.xlsx") -> set:
    workbook = load_workbook(excel_file_path)
    not_keywords_sheet = 'Training wav'
    keywords = set()
    for sheet_name in workbook.sheetnames:
        if sheet_name == not_keywords_sheet:
            continue
        sheet = workbook[sheet_name]
        for row in range(2,32):
            keywords.add(sheet[f"B{row}"].value)
    return keywords

# note: only works with dict structure, this is used for whisper prompt
def get_keywords_from_dict(key_dict:Dict[str, Dict[str, Tuple[int:str]]]) -> set:
    keywords = set()
    for value in key_dict.values():
        for key in value:
            keywords.add(key)
    return keywords

# note: only works with certain excel files structure
# structure, keyword_num for sharing between languages
"""
keyword_dict[language][keyword_string] -> (keyword_num, explanation)
num_dict[language][keyword_num] -> (keyword_string, explanation)
"""
# main excel opener, when this is called, other function that opens excel file is not required
def get_keywords_dictionary(excel_file_path:str = "Knowledge Dataset.xlsx") -> Tuple[Dict[str, Dict[str, Tuple[int:str]]], Dict[str, Dict[int, Tuple[str:str]]]]:
    workbook = load_workbook(excel_file_path)
    not_keywords_sheet = 'Training wav'
    name_lang = {'TW':LANGUAGES.TAIWANESE, 'US':LANGUAGES.ENGLISH, 'JP':LANGUAGES.JAPANESE, 'DE':LANGUAGES.GERMAN}
    keyword_dict = dict()
    num_dict = dict()
    for sheet_name in workbook.sheetnames:
        if sheet_name == not_keywords_sheet:
            continue
        sheet = workbook[sheet_name]
        language = name_lang[sheet_name[-2:]].value
        inner_keyword_dict = dict()
        inner_num_dict = dict()
        for row in range(2,32):
            explanation = sheet[f"C{row}"].value
            keyword_num = row-1
            keyword_string = sheet[f"B{row}"].value.strip()
            inner_keyword_dict[keyword_string] = (keyword_num, explanation)
            inner_num_dict[keyword_num] = (keyword_string, explanation)
        keyword_dict[language] = inner_keyword_dict
        num_dict[language] = inner_num_dict
    return keyword_dict, num_dict


class pattern_finder(object):
    def __init__(self, keyword_dict:Dict[str, Dict[str, Tuple[int:str]]]):
        self._automaton = dict()
        for language in LANGUAGES:
            language = language.value
            language_automaton = ahocorasick.Automaton()
            for pattern, (idx, explanation) in keyword_dict[language].items():
                language_automaton.add_word(pattern.lower(), idx)
            language_automaton.make_automaton()
            self._automaton[language] = language_automaton

    # returns List[int]: list of keyword's number
    def find_pattern(self, text:str, language:str) -> List[int]:
        matches = []
        for end_pos, pattern_id in self._automaton[language].iter(text.lower()):
            matches.append(pattern_id)
            # comment: below for finding position
            # match = patterns[pattern_id]
            # matches.append((match, end_pos - len(match) + 1, end_pos))
        return matches # TSMC requirement: don't remove duplicates
        return self._remove_duplicates_sorted(matches)
    
    def _remove_duplicates_sorted(self, lst) -> List[int]:
        seen = set()
        return [x for x in lst if not (x in seen or seen.add(x))]


# TSMC TASK 2: add explanation for keywords
class explainer(object):
    keyword_lang = {LANGUAGES.ENGLISH.value: "Keywords:", LANGUAGES.TAIWANESE.value: "關鍵字：",
                    LANGUAGES.JAPANESE.value: "キーワード:", LANGUAGES.GERMAN.value: "Schlüsselwörter:"}

    def __init__(self, keyword_num:dict):
        self._keyword_num = keyword_num

    def _get_keyword_explanation(self, keyword_number:int, language:str) -> str:
        keyword_explanation =  self._keyword_num[language][keyword_number]
        return ": ".join(keyword_explanation)

    # TSMC TASK 2:
    def explain_text(self, text:str, language:str, pattern_numbers:List[int] = None) -> str:
        # skip function if there's no pattern
        if not pattern_numbers:
            return text
        
        result = [text, explainer.keyword_lang[language]]
        for pattern_number in pattern_numbers:
            result.append(self._get_keyword_explanation(pattern_number, language))
        return "\n".join(result)


class text_translator(object):
    _translate_deepl_source_language_code = {LANGUAGES.ENGLISH.value:"EN", LANGUAGES.TAIWANESE.value:"ZH", LANGUAGES.JAPANESE.value:"JA", LANGUAGES.GERMAN.value:"DE"}
    _target_deepl_language_code = {LANGUAGES.ENGLISH.value:"EN-US", LANGUAGES.TAIWANESE.value:"ZH", LANGUAGES.JAPANESE.value:"JA", LANGUAGES.GERMAN.value:"DE"}
    def __init__(self, deepl_client):
        self._deepl_client = deepl_client
        self._init_glossary_name_id_dict()
    
    def _init_glossary_name_id_dict(self):
        glossary_name_id_dict = dict()
        for glossary_info in self._deepl_client.list_glossaries():
            glossary_name_id_dict[glossary_info.name] = glossary_info.glossary_id
        self._glossary_name_id_dict = glossary_name_id_dict

    def translate_text(self, text:str, source_language:str, target_language:str) -> str:
        if source_language == target_language:
            return text
        try:
            result = self.translate_deepl(text, source_language, target_language)
            if target_language == LANGUAGES.TAIWANESE.value:
                result = self.translate_google(result, 'zh-CN', target_language)
        except:
            # print("Deepl error, moving to google")
            result = self.translate_google(text, source_language, target_language)
        if result == "":
            result = text
        return result

    def translate_deepl(self, text:str, source_language:str, target_language:str) -> str:
        # normally at this point source and target won't be the same, but just in case
        if source_language == target_language:
            return text
        glossary_name = "_".join([source_language, target_language])
        glossary_id = self._glossary_name_id_dict[glossary_name]
        result = self._deepl_client.translate_text(
            text,  # Text to translate
            source_lang = text_translator._translate_deepl_source_language_code[source_language],  # Source language
            target_lang = text_translator._target_deepl_language_code[target_language],  # Target language
            glossary = glossary_id  # Use glossary
        )
        return result.text
    
    def translate_google(self, text:str, sl:str = "auto", tl:str = "zh-TW"):
        """On failure: return empty string
            On success: return translated string
        """
        # language codes: 'en', 'zh-TW', 'ja', 'de'
        if sl == "tw":
            sl = "zh-TW"
        if tl == "tw":
            tl = "zh-TW"
        if sl == tl:
            return text
        url = "https://translate.googleapis.com/translate_a/single"
        params = {
            "client": "gtx",
            "sl": sl,  # source language (auto-detect)
            "tl": tl,    # target language (English)
            "dt": "t",     # data type (translated text)
            "q": text     # the text to translate
        }
        response = requests.get(url, params=params)
        if response.status_code == 200:
            result = response.json()[0]
            sentences = []
            for res in result:
                sentences.append(res[0])
            translation = " ".join(sentences)
            return translation
        else:
            return ""

    # THIS FUNCTION SHOULD NOT BE CALLED BY ANYONE OTHER THAN THE API HOLDER
    # THIS FUNCTION SHOULD BE CALLED "NO" MORE THAN ONCE, EXCEPT WHEN REMAKING GLOSSARY
    def __DEEPL_MAKE_GLOSSARY__(self, num_dict:Dict[str, Dict[int, Tuple[str:str]]]):
        language_code = {LANGUAGES.ENGLISH.value:"EN", LANGUAGES.TAIWANESE.value:"ZH", LANGUAGES.JAPANESE.value:"JA", LANGUAGES.GERMAN.value:"DE"}
        for source_language in LANGUAGES:
            for target_language in LANGUAGES:
                if source_language == target_language:
                    continue
                source_language = source_language.value
                target_language = target_language.value
                glossary_name = "_".join([source_language,target_language])
                # uses keyword_number, since i'm not sure how the dictionary may be sorted
                # this function assumes both dictionary contain same amount of key and one to one correspondence
                # faster implementation uses both source and target, speeds up by around 2x
                source_dict = num_dict[source_language]
                target_dict = num_dict[target_language]
                entries = dict()
                for keyword_number in range(1,1+len(num_dict[source_language])):
                    entries[source_dict[keyword_number][0]] = target_dict[keyword_number][0]
                glossary = self._deepl_client.create_glossary(
                    name = glossary_name,  # Glossary name
                    source_lang = language_code[source_language],   # Source language
                    target_lang = language_code[target_language],   # Target language
                    entries = entries     # Dictionary of terms
                )

                
class meeting_translator(object):

    def __init__(self, openai_client, deepl_client):
        self._keyword_dict, self._num_dict = get_keywords_dictionary()
        self._STT_model = STT(openai_client, keywords=get_keywords_from_dict(self._keyword_dict))
        self._language_detector = lang_detector()
        self._keyword_finder = pattern_finder(self._keyword_dict)
        self._text_language_changer = text_translator(deepl_client)
        self._keyword_explainer = explainer(self._num_dict)
        self._transcribed_text = ""

    def translate_by_audio_path(self, audio_file_path:str, target_languages:Union[str,List[str]]) -> Union[str,List[str]]:
        audio_file = open(audio_file_path, "rb")
        return self.translate_by_audio(audio_file, target_languages)
    
    def translate_by_audio(self, audio_file, target_languages:str|List[str]=LANGUAGES.TAIWANESE.value) -> str|List[str]:
        self._transcribed_text = self._STT_model.transcript(audio_file)
        source_language = self._language_detector.detect_language_text(self._transcribed_text)
        if isinstance(target_languages, str):
            return self.translate_by_text(self._transcribed_text, source_language, target_languages)
        else:
            return self.translate_by_text_multi_language(self._transcribed_text, source_language, target_languages)
    
    def translate_by_text(self, text:str, source_language:str, target_language:str, keyword_nums:List[int]=None) -> str:
        # no need to check source-target language, since this also adds explained text
        if keyword_nums == None:
            keyword_nums = self._keyword_finder.find_pattern(text, source_language)
        translated_text = self._text_language_changer.translate_text(text, source_language, target_language)
        return translated_text
        # explained_text = self._keyword_explainer.explain_text(translated_text, target_language, keyword_nums)
        # return explained_text
    
    # BETA function, deepl may fail
    def translate_by_text_multi_language(self, text:str, source_language:str, target_languages:List[str], keyword_nums:List[int]=None) -> List[str]:
        if keyword_nums == None:
            keyword_nums = self._keyword_finder.find_pattern(text, source_language)
        results = []
        for target_language in target_languages:
            results.append(self.translate_by_text(text, source_language, target_language, keyword_nums))
        return results
    
    def _get_last_transcribed(self) -> str:
        return self._transcribed_text


def main():
    openai_client = OpenAI(api_key=OpenAI_API_KEY)
    deepl_client = deepl.Translator(DEEPL_API_KEY)
    translator_meeting = meeting_translator(openai_client, deepl_client)

    # single language usage
    TARGET_LANGUAGE = LANGUAGES.TAIWANESE.value
    
    folder_name = "text"
    if os.path.exists(folder_name):
        shutil.rmtree(folder_name)
    os.makedirs(folder_name)
    
    for i in range(1,19):
        # part_time_start = time.time()
        audio_file_path = f"upload_output_segments/segment_{i}.wav"
        
        output_text_path = f"text/part{i}.txt"
        output_text = translator_meeting.translate_by_audio_path(audio_file_path, TARGET_LANGUAGE)
        
        with open(output_text_path, "w", encoding="utf-8") as text_file:
            text_file.write(output_text)

        # Debugging purposes
        transcribed_text = translator_meeting._get_last_transcribed()
        transcribe_text_path = f"text/whisper_transcript_part{i}.txt"
        with open(transcribe_text_path, "w", encoding="utf-8") as text_file:
            text_file.write(transcribed_text)
    
    # # multi-language usage
    # TARGET_LANGUAGE = [LANGUAGES.TAIWANESE.value, LANGUAGES.ENGLISH.value, LANGUAGES.JAPANESE.value, LANGUAGES.GERMAN.value]
    # for i in range(1,14):
    #     # part_time_start = time.time()
    #     audio_file_path = f"audio/part{i}.wav"
    #     output_texts = translator_meeting.translate_by_audio_path(audio_file_path, TARGET_LANGUAGE)
    #     for language, output_text in zip(TARGET_LANGUAGE,output_texts):
    #         output_text_path = f"text/{language}_part{i}.txt"
    #         with open(output_text_path, "w", encoding="utf-8") as text_file:
    #             text_file.write(output_text)

    #     # Debugging purposes
    #     # transcribed_text = translator_meeting._get_last_trancribed()
    #     # transcribe_text_path = f"text/whisper_transcript_part{i}.txt"
    #     # with open(transcribe_text_path, "w", encoding="utf-8") as text_file:
    #     #     text_file.write(transcribed_text)
    # for i in range(19,20):
    #     result = translator_meeting._STT_model.transcript_by_path(f"segment/segment_{i}.wav")
    #     print(result)


if __name__ == "__main__":
    time_start = time.time()
    main()
    time_end = time.time()
    runtime = time_end-time_start
    print(f"Program runtime {runtime//60} minutes, {runtime%60} seconds")