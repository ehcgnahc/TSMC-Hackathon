import os
import time
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from google.cloud import speech
from google.oauth2 import service_account
from pydub import AudioSegment
from openpyxl import load_workbook

client_file = "google-credentials.json" # GCP服務帳戶金鑰
credentials = service_account.Credentials.from_service_account_file(client_file) # 讀取金鑰並建立Credentials (用於GCP API呼叫的認證物件)
client = speech.SpeechClient(credentials=credentials) # 建立SpeechClient (可直接呼叫STT API)
app = FastAPI() # 建立Application

# CORS(跨來源資源共享)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # 允許來自react的請求
    allow_credentials=True,
    allow_methods=["*"],  # 允許所有 HTTP 方法（GET、POST、PUT 等）
    allow_headers=["*"],  # 允許所有自訂標頭
)

# 音檔轉檔
def merge_audio_files(input_path, output_path):
    file_extension = input_path.split(".")[-1].lower()
    if file_extension in ("opus", "flac", "webm", "weba", "wav", "ogg", "m4a", "oga", "mid", "mid", "mp3", "aiff", "wma", "au"):
        audio = AudioSegment.from_file(input_path)
        audio = audio.set_channels(1)
        audio = audio.set_frame_rate(16000)
        audio = audio.set_sample_width(2)
        audio.export(output_path, format="wav")

# Keywords
def get_keywords(excel_file_path = "Knowledge Dataset.xlsx"):
    workbook = load_workbook(excel_file_path)
    not_keywords_sheet = 'Training wav'
    keywords = set()
    for sheet_name in workbook.sheetnames:
        if sheet_name == not_keywords_sheet:
            continue
        sheet = workbook[sheet_name]
        for row in range(2,32):
            keywords.add(sheet[f"B{row}"].value)
    return keywords

# 透過post接收音訊資料
@app.post("/api/upload")
async def upload_audio(file: UploadFile = File(...)):
    voice_path_webm = "voice_output.webm" # webm語音檔案路徑
    if os.path.exists(voice_path_webm):
        os.remove(voice_path_webm)
        print("Remove Old Voice File")

    with open(voice_path_webm, "wb") as f:
        content = await file.read()
        print(f"receive {len(content)} bytes") # 顯示接收到的資料長度
        f.write(content) # 寫入檔案
    
    voice_path_wav = "voice_output.wav"
    if os.path.exists(voice_path_wav):
        os.remove(voice_path_wav)
        print("Remove Old Voice File")

    merge_audio_files(voice_path_webm, voice_path_wav)


    config = speech.RecognitionConfig(
        encoding = speech.RecognitionConfig.AudioEncoding.LINEAR16,
        language_code = "en-US",
        alternative_language_codes = ["zh-TW", "ja-JP", "de-DE"],
        speech_contexts = [
            speech.SpeechContext(
                phrases=get_keywords(),
                boost = 10.0
            )],
        enable_automatic_punctuation = True,
    )
    
    start_time = time.time()
    
    # 檢查音檔格式
    input_audio = AudioSegment.from_wav(voice_path_wav)
    input_audio = input_audio.set_channels(1).set_frame_rate(16000).set_sample_width(2)

    print(f"音檔格式: {input_audio.channels} 通道, {input_audio.frame_rate} Hz, {input_audio.sample_width} 位元")
    
    with open(voice_path_wav, "rb") as f:
        content = f.read()

    audio = speech.RecognitionAudio(content=content)
    try:
        print("正在發送 STT 請求...")
        response = client.recognize(config=config, audio=audio)
        
        if not response.results:
            print("Warning:STT 請求成功但沒有回應")
            raise HTTPException(status_code=400, detail="No Speech Detected")
        else:
            print("Success:STT 請求成功！")
            for result in response.results:
                transcript = result.alternatives[0].transcript
                confidence = result.alternatives[0].confidence
                print(f"識別文字: {transcript}")
                print(f"信心指數: {confidence:.2f}")
                end_time = time.time()
                print(f"STT 處理時間: {end_time - start_time:.2f} 秒")
                return {"transcript": transcript, "confidence": confidence}
                
    except Exception as stt_error:
        error_message = f"STT Error: {str(stt_error)}"
        print(error_message)
        raise HTTPException(status_code=500, detail=error_message)
    print("Finish Receiving Audio")