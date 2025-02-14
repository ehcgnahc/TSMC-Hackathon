import subprocess
import json
import wave
import torch
import torchaudio
import noisereduce as nr
import soundfile as sf
from openpyxl import load_workbook
from pydub import AudioSegment
from demucs import pretrained
from demucs.apply import apply_model
import os


# Keywords
def get_keywords(excel_file_path = "Knowledge Dataset.xlsx"):
    workbook = load_workbook(excel_file_path)
    not_keywords_sheet = 'Training wav'
    keywords = set()
    for sheet_name in workbook.sheetnames:
        if sheet_name == not_keywords_sheet:
            continue
        sheet = workbook[sheet_name]
        for row in range(2,32):
            keywords.add(sheet[f"B{row}"].value)
    return keywords

# 音檔轉檔
def merge_audio_files(input_path, output_path):
    file_extension = input_path.split(".")[-1].lower()
    if file_extension in ("opus", "flac", "webm", "weba", "wav", "ogg", "m4a", "oga", "mid", "mid", "mp3", "aiff", "wma", "au"):
        audio = AudioSegment.from_file(input_path)
        audio = audio.set_channels(1)
        audio = audio.set_frame_rate(16000)
        audio = audio.set_sample_width(2)
        audio.export(output_path, format="wav")

# 音檔檢查
def get_audio_info(file_path):
    """
    使用 ffprobe 獲取音頻文件的詳細信息
    Args:
        file_path: 音頻文件路徑
    Returns:
        dict: 音頻文件的詳細信息
    """
    try:
        cmd = [
            'ffprobe',
            '-v', 'quiet',
            '-print_format', 'json',
            '-show_format',
            '-show_streams',
            file_path
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode == 0:
            info = json.loads(result.stdout)
            print(f"Audio file info for {file_path}:")
            print(json.dumps(info, indent=2, ensure_ascii=False))
            return info
        else:
            print(f"Error getting audio info: {result.stderr}")
            return None
            
    except Exception as e:
        print(f"Error running ffprobe: {str(e)}")
        return None

def save_as_wav(input_data, output_file):
    with wave.open(output_file, 'wb') as f:
        f.setnchannels(1)
        f.setsampwidth(2)
        f.setframerate(16000)
        f.writeframes(input_data)

def save_to_wav(input_data, output_file):
    with wave.open(output_file, 'wb') as f:
        f.setnchannels(2)
        f.setsampwidth(2)
        f.setframerate(16000)
        f.writeframes(input_data)

def reduce_noise(input_file, output_file):
    audio, sample_rate = sf.read(input_file)
    reduced_noise = nr.reduce_noise(y=audio, sr=sample_rate, prop_decrease=0.5, stationary=True)
    sf.write(output_file, reduced_noise, sample_rate)
    
def isolate_voice(input_file, output_dir):
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    model = pretrained.get_model("htdemucs")
    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    model.to(device)
    
    wav, sr = torchaudio.load(input_file)
    sources = apply_model(model, wav.unsqueeze(0), device=device, overlap=0.25, shifts=1, split=True)
    
    # sources的顺序为：[drums, bass, other, vocals]
    source_names = ["drums", "bass", "other", "vocals"]
    
    # 保存所有分离的音轨
    for source, name in zip(sources[0], source_names):
        output_path = os.path.join(output_dir, f"{name}.wav")
        torchaudio.save(output_path, source, sr)
    
    return os.path.join(output_dir, "vocals.wav")
