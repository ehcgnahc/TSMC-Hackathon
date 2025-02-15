import os
import time
import speech_recognition as sr
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openai import OpenAI
from openpyxl import load_workbook
from pydub import AudioSegment
from langdetect import detect_langs, DetectorFactory, detect
from Key import OpenAI_API_KEY

app = FastAPI()

DetectorFactory.seed = 0
POSSIBLE_LANGUAGES = ("en", "tw", "ja", "de") 
LANGUAGES = "en-US,ja-JP,de-DE,zh-TW"

# CORS(跨來源資源共享)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # 允許來自react的請求
    allow_credentials=True,
    allow_methods=["*"],  # 允許所有 HTTP 方法（GET、POST、PUT 等）
    allow_headers=["*"],  # 允許所有自訂標頭
)

class STT(object):
    def __init__(self, client, keywords=None):
        self.client = client
        if keywords == None:
            self.init_prompt = None
        else:
            self.init_prompt = self._make_init_prompt(keywords)

    def transcript(self, audio_file) -> str:
        result = self.client.audio.transcriptions.create(
            model="whisper-1", 
            file=audio_file,
            prompt=self.init_prompt
        )
        # print(result)
        return result.text

    def transcript_by_path(self, audio_file_path) -> str:
        audio_file = open(audio_file_path, "rb")
        return self.transcript(audio_file)

    def _make_init_prompt(self, keywords: list[str]) -> str:
        prompt =  "Recognize " + ", ".join(keywords) + ". Do not expand abbreviations."
        # print(prompt)
        return prompt

class lang_detector(object):
    
    def __init__(self):
        self.recognizer = sr.Recognizer()

    def detect_language_text(self, text:str):
        detected = detect_langs(text)
        filtered = [lang for lang in detected if lang.lang in POSSIBLE_LANGUAGES]
        if filtered:
            result = max(filtered, key=lambda l: l.prob).lang
        else:
            result = 'en'
        return result

# Keywords
def get_keywords(excel_file_path = "Knowledge Dataset.xlsx"):
    workbook = load_workbook(excel_file_path)
    not_keywords_sheet = 'Training wav'
    keywords = set()
    for sheet_name in workbook.sheetnames:
        if sheet_name == not_keywords_sheet:
            continue
        sheet = workbook[sheet_name]
        for row in range(2,32):
            keywords.add(sheet[f"B{row}"].value)
    return keywords

# 音檔轉檔
def merge_audio_files(input_path, output_path):
    file_extension = input_path.split(".")[-1].lower()
    if file_extension in ("opus", "flac", "webm", "weba", "wav", "ogg", "m4a", "oga", "mid", "mid", "mp3", "aiff", "wma", "au"):
        audio = AudioSegment.from_file(input_path)
        audio = audio.set_channels(1)
        audio = audio.set_frame_rate(16000)
        audio = audio.set_sample_width(2)
        audio.export(output_path, format="wav")
    
@app.post("/api/upload")
async def upload_audio(file: UploadFile = File(...)):
    voice_input_path = "voice_input." + file.filename.split(".")[-1]
    with open(voice_input_path, "wb") as f_out:
        content = await file.read()
        f_out.write(content)
    print(f"Received file {file.filename}, size = {len(content)} bytes")
    
    voice_path_wav = "voice_output.wav"
    try:
        if os.path.exists(voice_path_wav):
            os.remove(voice_path_wav)
            print("Remove Old Voice File")
        merge_audio_files(voice_input_path, voice_path_wav)
    except Exception as e:
        print(f"Error removing old voice file: {e}")
        raise HTTPException(status_code=500, detail="Error removing old voice file")
    
    time_start = time.time()
    
    client = OpenAI(api_key=OpenAI_API_KEY)
    stt_model = STT(client, keywords=get_keywords())
    my_language_detector = lang_detector()
       
    try:
        with open(voice_path_wav, "rb") as f:
            result = stt_model.transcript(f)
            language = my_language_detector.detect_language_text(result)
            print(f"Detected language: {language}")
            
            time_end = time.time()
            runtime = time_end-time_start
            print(f"Program runtime {runtime//60} minutes, {runtime%60} seconds")
            
            return {"transcript": result, "language": language, "runtime": runtime}
    except Exception as e:
        print(f"Error transcribing audio: {e}")
        raise HTTPException(status_code=500, detail="Error transcribing audio"+str(e))